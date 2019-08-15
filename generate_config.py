#!/usr/bin/env python
# -*- coding: utf-8 -*-

from xml.etree import ElementTree
from openpyxl import Workbook, styles

# Load Base Info
tree = ElementTree.parse("acdc.xml")
avamarConfig = tree.getroot()
gridConfig = avamarConfig.find(".//gridConfig")
systemName = gridConfig.get('systemName').split('.')[0]
localtime = avamarConfig.get('localtime').split('-')[0].replace('/', '')

# Open Excel
wb = Workbook()
ws = wb.active
ws.title = 'AvamarConfig'

# # Style
def form_cell(row, column, value):
    ws.merge_cells(start_row=row, end_row=row, start_column=column, end_column=6)
    target_cell = ws.cell(row=row, column=column, value=value)
    font = styles.Font(color='FFFFFF', bold=True)
    fill = styles.PatternFill(fill_type='solid', fgColor='0076CE')
    target_cell.fill = fill
    target_cell.font = font

line_count = 1
for e1 in avamarConfig:
    # Level 1
    # target_cell = ws.cell(row=line_count, column=1, value=e1.tag)
    form_cell(line_count, 1, e1.tag)
    line_count = line_count + 1

    for k in e1.attrib.keys():
        print(k, e1.attrib[k])
        ws.cell(row=line_count, column=2, value=k)
        ws.cell(row=line_count, column=3, value=e1.attrib[k])
        line_count = line_count + 1

    # Level 2
    for e2 in e1:
        # target_cell = ws.cell(row=line_count, column=2, value=e2.tag)
        form_cell(line_count, 2, e2.tag)
        line_count = line_count + 1

        for k in e2.attrib.keys():
            ws.cell(row=line_count, column=3, value=k)
            ws.cell(row=line_count, column=4, value=e2.attrib[k])
            line_count = line_count + 1

        # Level 3
        for e3 in e2:
            # target_cell = ws.cell(row=line_count, column=3, value=e3.tag)
            form_cell(line_count, 3, e3.tag)
            line_count = line_count + 1

            for k in e3.attrib.keys():
                ws.cell(row=line_count, column=4, value=k)
                ws.cell(row=line_count, column=5, value=e3.attrib[k])
                line_count = line_count + 1

            # Level 4
            for e4 in e3:
                # target_cell = ws.cell(row=line_count, column=4, value=e4.tag)
                form_cell(line_count, 4, e4.tag)
                line_count = line_count + 1

                for k in e4.attrib.keys():
                    ws.cell(row=line_count, column=5, value=k)
                    ws.cell(row=line_count, column=6, value=e4.attrib[k])
                    line_count = line_count + 1


wb.save(systemName + '_' + localtime + '.xlsx')
